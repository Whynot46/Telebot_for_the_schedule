import telebot
from telebot import types
from openpyxl import load_workbook
from datetime import date, datetime, timedelta
import calendar
from colorama import init
from colorama import Fore
from time import sleep
import os

os.system('cls')
init()

# подключение к таблицам
try:
    wb_timetable = load_workbook('./db/timetable.xlsx')
    numerator_sheet = wb_timetable['numerator_list']
    denominator_sheet = wb_timetable['denominator_list']
    print(Fore.WHITE+'Loading tables: '+Fore.GREEN+'ok')
except:
    print(Fore.WHITE +'Loading tables: '+Fore.RED+'ERROR')

# запуск бота Jarvis https://t.me/ASU_BDMS_11_Jarvis_bot
try:
    bot = telebot.TeleBot('API_KEY')
    bot_name = 'Jarvis'
    print(Fore.WHITE+'Launching the bot '+bot_name+' : '+Fore.GREEN+'ok\n')
except:
    print(Fore.WHITE + 'Launching the bot '+bot_name+' : ' + Fore.RED + 'ERROR\n')

@bot.message_handler(commands=['start'])
def start_message(message):
    global user_data
    user_data = '\nUsername: '+str(message.from_user.username)+'\nID: '+str(message.from_user.id)+'\nFirst name: '+str(message.from_user.first_name)+'\nLast_name: '+str(message.from_user.last_name)+'\n---------------------'
    write_user_data()
    #print(bot.get_me())

    markup = types.ReplyKeyboardMarkup(row_width=2)
    btn_commands = types.KeyboardButton('/Команды🤖')
    markup.add(btn_commands)
    bot.send_message(message.chat.id,
                     'Здравствуй, друг! Меня зовут Jarvis👾\n\n'
                     'Забыл какая завтра пара😴?\n'
                     'Никак не можешь вспомнить как зовут преподавателя🧐?\n'
                     'Хочешь позвать одногруппников на вечеринку,\nно не помнишь, какие их настоящие имена🫣?\n'
                     'Устал сидеть на паре\nи хочешь знать сколько осталось до конца🥴?\n'
                     'Хочешь узнать последние новости Агу🥳?\n\n'
                     'С этим и многим другим я постараюсь тебе помочь🙃',
                     reply_markup=markup)

@bot.message_handler(commands=['Команды🤖', 'help'])
def commands(message):
    markup = types.ReplyKeyboardMarkup(row_width=2)
    btn_links = types.KeyboardButton('/Ссылки↗️')
    btn_timetable = types.KeyboardButton('/Расписание🗓')
    btn_teachers = types.KeyboardButton('/Преподаватели👩‍🏫👨‍🏫')
    btn_studens = types.KeyboardButton('/Одногруппники🤪')
    btn_time = types.KeyboardButton('/Время🕒')
    btn_news = types.KeyboardButton('/Новости🪩📣🔈')
    btn_today = types.KeyboardButton('/Сегодня👀')
    markup.add(btn_links, btn_timetable, btn_teachers, btn_time, btn_studens, btn_news, btn_today)
    bot.send_message(message.chat.id, "Команды 🥸⁉️", reply_markup=markup)
    print(Fore.WHITE+'--> a new '+Fore.GREEN+'command'+Fore.WHITE+' has been received: '+Fore.RED+str(message.from_user.first_name)+' '+str(message.from_user.last_name)+'-->'+Fore.GREEN+message.text+Fore.RED+'<--'+str(datetime.now()))

@bot.message_handler(commands=['Ссылки↗️'])
def links(message):
    markup = types.InlineKeyboardMarkup()
    btn_asu_link = types.InlineKeyboardButton(text='Сайт АГУ', url='https://asu.edu.ru')
    btn_bdms_link = types.InlineKeyboardButton(text='Разговорная беседа', url='https://t.me/+Quwz3hi48c0wZGU6')
    btn_ads = types.InlineKeyboardButton(text='Объявления', url='https://t.me/+uAWhg9U1WeZkNjgy')
    btn_timetabs = types.InlineKeyboardButton(text='Расписание', url='https://raspisanie.asu.edu.ru/student/ДМС11')
    markup.add(btn_asu_link, btn_bdms_link, btn_ads, btn_timetabs)
    bot.send_message(message.chat.id, "Список полезных ссылок🧠\nНаша почта: bdms00011@gmail.com", reply_markup=markup)
    print(Fore.WHITE+'--> a new '+Fore.GREEN+'command'+Fore.WHITE+' has been received: '+Fore.RED+str(message.from_user.first_name)+' '+str(message.from_user.last_name)+'-->'+Fore.GREEN+message.text+Fore.RED+'<--'+str(datetime.now()))

@bot.message_handler(commands=['Преподаватели👩‍🏫👨‍🏫'])
def teachers(message):
    markup = types.InlineKeyboardMarkup()
    btn_krivyh_l_d = types.InlineKeyboardButton(text='Кривых Л.Д.', url='https://asu.edu.ru/staff/622-1-krivyh-ludmila-dmitrievna.html')
    btn_trofimova_a_n = types.InlineKeyboardButton(text='Трофимова А.Н.',url='https://asu.edu.ru/staff/1279-1-trofimova-anna-nikolaevna.html')
    btn_kanash_e_sh = types.InlineKeyboardButton(text='Канаш Э.Ш.', url='https://asu.edu.ru/staff/699-1-kanash-elvira-shavukatovna.html')
    btn_turin_a_o = types.InlineKeyboardButton(text='Тюрин А.О.', url='https://asu.edu.ru/staff/333-1-turin-aleksei-olegovich.html')
    btn_smirnov_a_p = types.InlineKeyboardButton(text='Смирнов А.П.', url='https://asu.edu.ru/staff/412-1-smirnov-aleksandr-petrovich.html')
    btn_kolomina_m_v = types.InlineKeyboardButton(text='Коломина М.В.', url='https://asu.edu.ru/staff/404-1-kolomina-marina-vladimirovna.html')
    btn_duhnov_i_n = types.InlineKeyboardButton(text='Духнов И.Н. ', url='https://asu.edu.ru/staff/1474-1-duhnov-ivan-nikolaevich.html')
    btn_turina_i_u = types.InlineKeyboardButton(text='Тюрина И.Ю.', url='https://asu.edu.ru/staff/605-1-turina-irina-urevna.html')
    btn_vostrikov_i_v = types.InlineKeyboardButton(text='Востриков И.В.', url='https://asu.edu.ru/staff/335-1-vostrikov-igor-vladimirovich.html')
    btn_ivashinenko_e_a = types.InlineKeyboardButton(text='Ивашиненко Е.А.', url='https://asu.edu.ru/staff/1620-1-ivashinenko-ekaterina-aleksandrovna.html')
    btn_chuikov_u_s = types.InlineKeyboardButton(text='Чуйков Ю.С.', url='https://asu.edu.ru/staff/105-1-chuikov-urii-sergeevich.html')
    btn_bibarsov_d_a = types.InlineKeyboardButton(text='Бибарсов Д.А.', url='https://asu.edu.ru/staff/1409-1-bibarsov-dmitrii-aleksandrovich.html')
    btn_ajmuhamedov_i_m = types.InlineKeyboardButton(text='Ажмухамедов И. М.', url='https://asu.edu.ru/staff/1145-1-ajmuhamedov-iskandar-maratovich.html')



    markup.add(btn_krivyh_l_d, btn_trofimova_a_n, btn_kanash_e_sh, btn_turin_a_o, btn_smirnov_a_p, btn_kolomina_m_v,
               btn_duhnov_i_n, btn_turina_i_u, btn_vostrikov_i_v,btn_ivashinenko_e_a, btn_chuikov_u_s, btn_bibarsov_d_a,  btn_ajmuhamedov_i_m)
    bot.send_message(message.chat.id, "Список преподавателей😻🫶", reply_markup=markup)
    print(Fore.WHITE + '--> a new ' + Fore.GREEN + 'command' + Fore.WHITE + ' has been received: ' + Fore.RED + str(message.from_user.first_name) + ' ' + str(message.from_user.last_name) + '-->' + Fore.GREEN + message.text + Fore.RED + '<--' + str(datetime.now()))

@bot.message_handler(commands=['Одногруппники🤪'])
def stud_list_image(message):
    markup = types.ReplyKeyboardMarkup(row_width=2)
    btn_students_list_1 = types.KeyboardButton('/Список_1')
    btn_students_list_2 = types.KeyboardButton('/Список_2')
    btn_students_image = types.KeyboardButton('/Фото')
    btn_commands = types.KeyboardButton('/Команды🤖')
    markup.add(btn_students_list_1, btn_students_list_2, btn_students_image, btn_commands)
    bot.send_message(message.chat.id, "В каком формате показать список одногруппников🤔?", reply_markup=markup)
    print(Fore.WHITE + '--> a new ' + Fore.GREEN + 'command' + Fore.WHITE + ' has been received: ' + Fore.RED + str(message.from_user.first_name) + ' ' + str(message.from_user.last_name) + '-->' + Fore.GREEN + message.text + Fore.RED + '<--' + str(datetime.now()))

@bot.message_handler(commands=['Список_1'])
def stud_list_show_1(message):
    try:
        students_list_1= open('./db/students_1.txt', 'r', encoding='utf-8')
        bot.send_message(message.chat.id, students_list_1.read())
        students_list_1.close()
        print(Fore.WHITE+'--> a new '+Fore.GREEN+'command'+Fore.WHITE+' has been received: '+Fore.RED+str(message.from_user.first_name)+' '+str(message.from_user.last_name)+'-->'+Fore.GREEN+message.text+Fore.RED+'<--'+str(datetime.now()))
    except:
        print(Fore.RED + 'ERROR : stud_list_show')

@bot.message_handler(commands=['Список_2'])
def stud_list_show_2(message):
    try:
        students_list_2 = open('./db/students_2.txt', 'r', encoding='utf-8')
        bot.send_message(message.chat.id, students_list_2.read())
        students_list_2.close()
        print(Fore.WHITE+'--> a new '+Fore.GREEN+'command'+Fore.WHITE+' has been received: '+Fore.RED+str(message.from_user.first_name)+' '+str(message.from_user.last_name)+'-->'+Fore.GREEN+message.text+Fore.RED+'<--'+str(datetime.now()))
    except:
        print(Fore.RED + 'ERROR : stud_list_show_2')

@bot.message_handler(commands=['Фото'])
def stud_image_show(message):
    try:
        bot.send_photo(message.chat.id, photo=open('./image/stud_list.png', 'rb'))
        bot.send_photo(message.chat.id, photo=open('./image/groups.jpg', 'rb'))
        print(Fore.WHITE+'--> a new '+Fore.GREEN+'command'+Fore.WHITE+' has been received: '+Fore.RED+str(message.from_user.first_name)+' '+str(message.from_user.last_name)+'-->'+Fore.GREEN+message.text+Fore.RED+'<--'+str(datetime.now()))
    except:
        print(Fore.RED+'ERROR : stud_image_show')

@bot.message_handler(commands=['Сегодня👀'])
def today(message):
    global parity
    my_date = date.today()
    wk = calendar.day_name[my_date.weekday()]
    try:
        now = datetime.now()
        sep = datetime(now.year if now.month >= 9 else now.year - 1, 9, 1)
        d1 = sep - timedelta(days=sep.weekday())
        d2 = now - timedelta(days=now.weekday())
        parity = ((d2 - d1).days // 7) % 2
    except: print(Fore.RED+'ERROR : it is impossible to determine the time')

    bot.send_message(message.chat.id, "Расписание на сегодня🥱")
    print(Fore.WHITE+'--> a new '+Fore.GREEN+'command'+Fore.WHITE+' has been received: '+Fore.RED+str(message.from_user.first_name)+' '+str(message.from_user.last_name)+'-->'+Fore.GREEN+message.text+Fore.RED+'<--'+str(datetime.now()))
    if wk == 'Sunday':
        bot.send_message(message.chat.id, "Сегодня воскресенье🥱")
    else:
        f = globals()[wk]
        f(message)

@bot.message_handler(commands=['Расписание🗓'])
def timetable(message):
    global parity
    try:
        now = datetime.now()
        sep = datetime(now.year if now.month >= 9 else now.year - 1, 9, 1)
        d1 = sep - timedelta(days=sep.weekday())
        d2 = now - timedelta(days=now.weekday())
        parity = ((d2 - d1).days // 7) % 2
    except: print(Fore.RED+'ERROR : it is impossible to determine the time')

    markup = types.ReplyKeyboardMarkup(row_width=2)
    btn_monday = types.KeyboardButton('/Понедельник1️⃣')
    btn_tuesday = types.KeyboardButton('/Вторник2️⃣')
    btn_wednesday = types.KeyboardButton('/Среда3️⃣')
    btn_thursday = types.KeyboardButton('/Четверг4️⃣')
    btn_friday = types.KeyboardButton('/Пятница5️⃣')
    btn_saturday = types.KeyboardButton('/Суббота6️⃣')
    btn_commands = types.KeyboardButton('/Команды🤖')
    markup.add(btn_monday, btn_tuesday, btn_wednesday, btn_thursday, btn_friday, btn_saturday, btn_commands)
    bot.send_message(message.chat.id, "Расписание🤯\nПосмотреть полное расписание:  /timetable\nПросмотреть время начала/окончаний занятий:  /time", reply_markup=markup)
    print(Fore.WHITE+'--> a new '+Fore.GREEN+'command'+Fore.WHITE+' has been received: '+Fore.RED+str(message.from_user.first_name)+' '+str(message.from_user.last_name)+'-->'+Fore.GREEN+message.text+Fore.RED+'<--'+str(datetime.now()))

@bot.message_handler(commands=['timetable'])
def timetable_image(message):
    bot.send_sticker(message.chat.id, 'CAACAgIAAxkBAAEFvAtjEPrdsgcpnKgW5GNQhk58-hRYHAAChBYAAtL60UushSNGNq8wTCkE')
    sleep(5)
    bot.send_message(message.chat.id, 'Ахахаха, ладно, шучу, держи😂')
    bot.send_photo(message.chat.id, photo=open('./image/timetable.PNG', 'rb'))
    print(Fore.WHITE+'--> a new '+Fore.GREEN+'command'+Fore.WHITE+' has been received: '+Fore.RED+str(message.from_user.first_name)+' '+str(message.from_user.last_name)+'-->'+Fore.GREEN+message.text+Fore.RED+'<--'+str(datetime.now()))

@bot.message_handler(commands=['time'])
def timetable_image(message):
    bot.send_photo(message.chat.id, photo=open('./image/time.png', 'rb'))
    print(Fore.WHITE+'--> a new '+Fore.GREEN+'command'+Fore.WHITE+' has been received: '+Fore.RED+str(message.from_user.first_name)+' '+str(message.from_user.last_name)+'-->'+Fore.GREEN+message.text+Fore.RED+'<--'+str(datetime.now()))

@bot.message_handler(commands=['Понедельник1️⃣'])
def Monday(message):
    print(Fore.WHITE+'--> a new '+Fore.GREEN+'command'+Fore.WHITE+' has been received: '+Fore.RED+str(message.from_user.first_name)+' '+str(message.from_user.last_name)+'-->'+Fore.GREEN+message.text+Fore.RED+'<--'+str(datetime.now()))
    if parity:  # нечетная неделя
        bot.send_message(message.chat.id, denominator_sheet.cell(row=2, column=1).value)
    else:  # четная неделя
        bot.send_message(message.chat.id, numerator_sheet.cell(row=2, column=1).value)

@bot.message_handler(commands=['Вторник2️⃣'])
def Tuesday(message):
    print(Fore.WHITE+'--> a new '+Fore.GREEN+'command'+Fore.WHITE+' has been received: '+Fore.RED+str(message.from_user.first_name)+' '+str(message.from_user.last_name)+'-->'+Fore.GREEN+message.text+Fore.RED+'<--'+str(datetime.now()))
    if parity:  # нечетная неделя
        bot.send_message(message.chat.id, denominator_sheet.cell(row=2, column=2).value)
    else:  # четная неделя
        bot.send_message(message.chat.id, numerator_sheet.cell(row=2, column=2).value)

@bot.message_handler(commands=['Среда3️⃣'])
def Wednesday(message):
    print(Fore.WHITE+'--> a new '+Fore.GREEN+'command'+Fore.WHITE+' has been received: '+Fore.RED+str(message.from_user.first_name)+' '+str(message.from_user.last_name)+'-->'+Fore.GREEN+message.text+Fore.RED+'<--'+str(datetime.now()))
    if parity:  # нечетная неделя
        bot.send_message(message.chat.id, denominator_sheet.cell(row=2, column=3).value)
    else:  # четная неделя
        bot.send_message(message.chat.id, numerator_sheet.cell(row=2, column=3).value)

@bot.message_handler(commands=['Четверг4️⃣'])
def Thursday(message):
    print(Fore.WHITE+'--> a new '+Fore.GREEN+'command'+Fore.WHITE+' has been received: '+Fore.RED+str(message.from_user.first_name)+' '+str(message.from_user.last_name)+'-->'+Fore.GREEN+message.text+Fore.RED+'<--'+str(datetime.now()))
    if parity:  # нечетная неделя
        bot.send_message(message.chat.id, denominator_sheet.cell(row=2, column=4).value)
    else:  # четная неделя
        bot.send_message(message.chat.id, numerator_sheet.cell(row=2, column=4).value)

@bot.message_handler(commands=['Пятница5️⃣'])
def Friday(message):
    print(Fore.WHITE+'--> a new '+Fore.GREEN+'command'+Fore.WHITE+' has been received: '+Fore.RED+str(message.from_user.first_name)+' '+str(message.from_user.last_name)+'-->'+Fore.GREEN+message.text+Fore.RED+'<--'+str(datetime.now()))
    if parity:  # нечетная неделя
        bot.send_message(message.chat.id, denominator_sheet.cell(row=2, column=5).value)
    else:  # четная неделя
        bot.send_message(message.chat.id, numerator_sheet.cell(row=2, column=5).value)

@bot.message_handler(commands=['Суббота6️⃣'])
def Saturday(message):
    print(Fore.WHITE+'--> a new '+Fore.GREEN+'command'+Fore.WHITE+' has been received: '+Fore.RED+str(message.from_user.first_name)+' '+str(message.from_user.last_name)+'-->'+Fore.GREEN+message.text+Fore.RED+'<--'+str(datetime.now()))
    if parity:  # нечетная неделя
        bot.send_message(message.chat.id, denominator_sheet.cell(row=2, column=6).value)
    else:  # четная неделя
        bot.send_message(message.chat.id, numerator_sheet.cell(row=2, column=6).value)

@bot.message_handler(commands=['Время🕒'])
def time(message):
    print(Fore.WHITE+'--> a new '+Fore.GREEN+'command'+Fore.WHITE+' has been received: '+Fore.RED+str(message.from_user.first_name)+' '+str(message.from_user.last_name)+'-->'+Fore.GREEN+message.text+Fore.RED+'<--'+str(datetime.now()))
    bot.send_message(message.chat.id, ('😴В разработке😴'))

@bot.message_handler(commands=['Новости🪩📣🔈'])
def news(message):
    print(Fore.WHITE+'--> a new '+Fore.GREEN+'command'+Fore.WHITE+' has been received: '+Fore.RED+str(message.from_user.first_name)+' '+str(message.from_user.last_name)+'-->'+Fore.GREEN+message.text+Fore.RED+'<--'+str(datetime.now()))
    bot.send_message(message.chat.id, ('😴В разработке😴'))

#запись сообщений пользователей
@bot.message_handler(content_types=['text'])
def send_text(message):
    try:
        users_messages_txt = open('./log/Users_messages.txt', 'a')
        users_messages_txt.write('\n'+str(message.from_user.first_name)+' '+str(message.from_user.last_name)+'-->'+message.text+'<--'+str(datetime.now())+'\n---------------------')
        bot.send_message(message.chat.id, ('Я непременно передам ваши слова своему создателю😉'))
        print(Fore.WHITE+'--> a new '+Fore.RED+'message'+Fore.WHITE+' has been received: '+Fore.RED+str(message.from_user.first_name)+' '+str(message.from_user.last_name)+'-->'+Fore.GREEN+message.text+Fore.RED+'<--'+str(datetime.now()))
    except UnicodeEncodeError:
        print(Fore.WHITE + '--> a new '+Fore.RED+'message'+Fore.WHITE+' has been received: ' + Fore.RED + 'UnicodeEncodeError')
        bot.send_message(message.chat.id, ('Мои нейроны пока что не могу распознавать эмодзи😭\nПожалуйста, введи коректное сообщение🥺🙏'))
    except:
        print(Fore.WHITE + '--> a new '+Fore.RED+'message'+Fore.WHITE+' has been received: ' + Fore.RED + 'Error')
        bot.send_message(message.chat.id, ('Мои нейроны не смогли обработать твоё сообщение😭\nПожалуйста, введи коректное сообщение🥺🙏'))

#запись данных о пользователях
def write_user_data():
        users_data_txt = open('./log/Users_data.txt', 'a')
        users_data_txt.write(user_data)
        print(Fore.GREEN+'--> a new user is connected')
        users_data_txt.close()


if __name__ == "__main__":
    bot.polling(none_stop=True, interval=0, timeout=20)
